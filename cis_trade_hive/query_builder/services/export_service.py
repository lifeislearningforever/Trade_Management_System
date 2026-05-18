"""
Export Service

Single responsibility: convert result rows to download formats.
No network I/O, no DB calls, no business logic.
"""

import csv
import io
import json
import logging
from datetime import datetime
from typing import List, Dict, Generator, Optional

logger = logging.getLogger(__name__)


class ExportService:
    """Converts query result rows into various download formats."""

    MAX_PDF_ROWS = 2_000

    # PDF layout constants
    _PDF_MARGIN = 36        # points (0.5 inch)
    _PDF_HEADER_BLUE = (31, 78, 121)
    _PDF_ALT_ROW = (240, 245, 250)
    _PDF_FONT_HEADER = 9
    _PDF_FONT_BODY = 8
    _PDF_MIN_COL_WIDTH = 50
    _PDF_MAX_COL_WIDTH = 180

    def to_csv_streaming(self, rows: List[Dict], columns: List[str] = None) -> Generator:
        """Generator suitable for Django StreamingHttpResponse."""
        cols = columns or (list(rows[0].keys()) if rows else [])
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(cols)
        yield buf.getvalue()
        for row in rows:
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow([row.get(c, '') for c in cols])
            yield buf.getvalue()

    def to_excel_bytes(
        self,
        rows: List[Dict],
        columns: List[str] = None,
        sheet_name: str = 'Results',
    ) -> bytes:
        """Return Excel workbook as bytes. Requires openpyxl."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export: pip install openpyxl")

        cols = columns or (list(rows[0].keys()) if rows else [])
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet(title=sheet_name[:31])

        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_row = []
        for col in cols:
            cell = openpyxl.cell.WriteOnlyCell(ws, value=str(col))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            header_row.append(cell)
        ws.append(header_row)

        for row in rows:
            ws.append([row.get(c, '') for c in cols])

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def to_json_streaming(self, rows: List[Dict]) -> Generator:
        """Generator that streams a JSON array."""
        yield '[\n'
        for i, row in enumerate(rows):
            suffix = ',\n' if i < len(rows) - 1 else '\n'
            yield json.dumps(row, default=str) + suffix
        yield ']\n'

    def to_pdf_bytes(
        self,
        rows: List[Dict],
        columns: List[str] = None,
        title: str = 'Query Results',
        subtitle: Optional[str] = None,
        generated_by: Optional[str] = None,
    ) -> bytes:
        """
        Return a PDF document as bytes.

        Layout: landscape A4, styled header, alternating row shading,
        footer with page number + generated timestamp, truncated at MAX_PDF_ROWS.
        Requires reportlab.
        """
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph,
                Spacer, HRFlowable,
            )
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        except ImportError:
            raise ImportError("reportlab is required for PDF export: pip install reportlab")

        cols = columns or (list(rows[0].keys()) if rows else [])
        data_rows = rows[:self.MAX_PDF_ROWS]
        truncated = len(rows) > self.MAX_PDF_ROWS

        buf = io.BytesIO()
        page_size = landscape(A4)
        page_w, page_h = page_size

        doc = SimpleDocTemplate(
            buf,
            pagesize=page_size,
            leftMargin=self._PDF_MARGIN,
            rightMargin=self._PDF_MARGIN,
            topMargin=self._PDF_MARGIN + 10,
            bottomMargin=self._PDF_MARGIN + 20,
            title=title,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'QBTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#1F4E79'),
            spaceAfter=4,
        )
        sub_style = ParagraphStyle(
            'QBSub',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.HexColor('#555555'),
            spaceAfter=2,
        )
        footer_style = ParagraphStyle(
            'QBFooter',
            parent=styles['Normal'],
            fontSize=7,
            textColor=colors.grey,
            alignment=TA_RIGHT,
        )

        # ---- header block ----
        generated_at = datetime.now().strftime('%Y-%m-%d %H:%M')
        meta_parts = [f"Generated: {generated_at}"]
        if generated_by:
            meta_parts.insert(0, f"By: {generated_by}")
        if truncated:
            meta_parts.append(f"⚠ Showing first {self.MAX_PDF_ROWS:,} of {len(rows):,} rows")

        story = [
            Paragraph(title, title_style),
        ]
        if subtitle:
            story.append(Paragraph(subtitle, sub_style))
        story.append(Paragraph(' | '.join(meta_parts), sub_style))
        story.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#1F4E79')))
        story.append(Spacer(1, 8))

        # ---- table ----
        usable_w = page_w - 2 * self._PDF_MARGIN
        col_count = len(cols)
        # Distribute width proportionally (clamped)
        raw_w = usable_w / max(col_count, 1)
        col_w = max(self._PDF_MIN_COL_WIDTH, min(self._PDF_MAX_COL_WIDTH, raw_w))
        col_widths = [col_w] * col_count

        header_row_data = [str(c) for c in cols]
        table_data = [header_row_data]
        for row in data_rows:
            table_data.append([
                str(row.get(c, '') if row.get(c) is not None else '')
                for c in cols
            ])

        hdr_blue = colors.Color(
            self._PDF_HEADER_BLUE[0] / 255,
            self._PDF_HEADER_BLUE[1] / 255,
            self._PDF_HEADER_BLUE[2] / 255,
        )
        alt_color = colors.Color(
            self._PDF_ALT_ROW[0] / 255,
            self._PDF_ALT_ROW[1] / 255,
            self._PDF_ALT_ROW[2] / 255,
        )

        tbl_style = [
            # Header
            ('BACKGROUND',   (0, 0), (-1, 0),  hdr_blue),
            ('TEXTCOLOR',    (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',     (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0), (-1, 0),  self._PDF_FONT_HEADER),
            ('ALIGN',        (0, 0), (-1, 0),  'CENTER'),
            ('BOTTOMPADDING',(0, 0), (-1, 0),  6),
            ('TOPPADDING',   (0, 0), (-1, 0),  6),
            # Body
            ('FONTNAME',     (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE',     (0, 1), (-1, -1), self._PDF_FONT_BODY),
            ('ALIGN',        (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',   (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING',(0, 1), (-1, -1), 3),
            # Grid
            ('GRID',         (0, 0), (-1, -1), 0.4, colors.HexColor('#C5D0DC')),
            ('LINEBELOW',    (0, 0), (-1, 0),  1.5, hdr_blue),
        ]
        # Alternating row shading
        for row_i in range(1, len(table_data)):
            if row_i % 2 == 0:
                tbl_style.append(('BACKGROUND', (0, row_i), (-1, row_i), alt_color))

        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle(tbl_style))
        story.append(tbl)

        # ---- page number footer via onLaterPages ----
        def _add_footer(canvas, doc):
            canvas.saveState()
            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(colors.grey)
            txt = f"Page {doc.page}  |  {title}  |  {generated_at}"
            canvas.drawRightString(page_w - self._PDF_MARGIN, 18, txt)
            canvas.restoreState()

        doc.build(story, onFirstPage=_add_footer, onLaterPages=_add_footer)
        return buf.getvalue()


export_service = ExportService()
